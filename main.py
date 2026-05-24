"""Точка входа парсера МФО (Selenium).

Запуск:
    python main.py chrome         # разовый запуск Chrome с remote debugging :9222
    python main.py run            # боевой: открыть ссылку, оставить заявку, ловить SMS + звонки
    python main.py demo           # демо без браузера и без API SMS
    python main.py show           # показать таблицу events / calls

Цепочка:
    1. SmsProvider — даёт номер (webhook = Android, manual = ручной ввод).
    2. Selenium (CDP attach или undetected) — открывает партнёрскую ссылку,
       проходит лендинг МФО, заполняет анкету фейковыми данными и шлёт SMS-код.
    3. Параллельно слушаем входящие SMS и звонки (Android webhook или ручной ввод).
    4. Когда SMS приходит — парсер вытаскивает код, пишет SMS в таблицу
       и сам вписывает код в поле «Код подтверждения».
    5. Каждое событие пишется в SQLite + CSV + Excel.
"""
from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
import threading
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from loguru import logger

from config import settings
from src.ai.classifier import AIClassifier
from src.calls import build_call_provider
from src.form.browser import (
    CDP_PORT,
    ensure_cdp_chrome_running,
    kill_cdp_chrome,
    open_browser,
    resolve_user_agent,
)
from src.form.fake_data import generate_persona
from src.form.filler import fill_and_submit_sync
from src.redirect.tracker import extract_urls, resolve_final_url
from src.sms import build_provider
from src.storage.logger import EventStorage


CALL_ENTRY_RX = re.compile(r"^\s*(?:call\s*[\|:]\s*)?(\+?\d[\d\s().-]{8,}\d)\s*$", re.IGNORECASE)
OTP_HINT_RX = re.compile(
    r"web[-\s]?zaim|webzaim|web-zaim\.ru|веб[-\s]?займ|"
    r"\b(код|code|otp|pin)\b|парол|password|подтвержд|confirm",
    re.IGNORECASE,
)
WEBZAIM_HINT_RX = re.compile(
    r"web[-\s]?zaim|webzaim|web-zaim\.ru|веб[-\s]?займ",
    re.IGNORECASE,
)
GENERIC_SMS_SENDER_RX = re.compile(
    r"^\s*(?:sim\s*\d+|sms|iphone|forward(?:ed)?\s*sms|unknown)\s*$",
    re.IGNORECASE,
)


def _domain_from_url(url: str | None) -> str:
    if not url:
        return ""
    parsed = urlparse(url)
    host = parsed.netloc or parsed.path.split("/", 1)[0]
    return host.removeprefix("www.").lower()


def infer_sender_alpha(sender: str, text: str, traces: list | None) -> str:
    """If Forward SMS only gives SIM slot, infer alpha-name from the SMS link."""
    sender = (sender or "").strip() or "unknown"
    if not GENERIC_SMS_SENDER_RX.match(sender):
        return sender

    for trace in traces or []:
        domain = _domain_from_url(getattr(trace, "original", None))
        if domain:
            logger.info(f"sender_alpha: {sender!r} заменён на домен ссылки {domain!r}")
            return domain
        domain = _domain_from_url(getattr(trace, "final", None))
        if domain:
            logger.info(f"sender_alpha: {sender!r} заменён на финальный домен {domain!r}")
            return domain

    urls = extract_urls(text)
    if urls:
        domain = _domain_from_url(urls[0])
        if domain:
            logger.info(f"sender_alpha: {sender!r} заменён на домен из текста {domain!r}")
            return domain
    return sender


async def cancel_listener_tasks(*tasks: asyncio.Task) -> None:
    for task in tasks:
        if not task.done():
            task.cancel()
    await asyncio.gather(*tasks, return_exceptions=True)


def parse_manual_call_entry(text: str) -> str | None:
    """Bare +7... or call|+7... in manual input means an incoming call."""
    match = CALL_ENTRY_RX.match(text or "")
    if not match:
        return None
    digits = re.sub(r"\D", "", match.group(1))
    if len(digits) < 10:
        return None
    if digits.startswith("8") and len(digits) == 11:
        digits = "7" + digits[1:]
    return "+" + digits if not match.group(1).strip().startswith("+") else "+" + digits


def setup_logging() -> None:
    logger.remove()
    logger.add(
        sys.stderr,
        level=settings.log_level,
        format="<green>{time:HH:mm:ss}</green> | <level>{level: <7}</level> | {message}",
    )
    logger.add(
        settings.output_dir / "run.log",
        level="DEBUG",
        rotation="5 MB",
        encoding="utf-8",
        enqueue=True,
    )


def cmd_check() -> int:
    """Проверка окружения перед боевым прогоном."""
    setup_logging()
    ok = True
    logger.info("Проверка окружения…")
    logger.info(f"  TRACKER_URL = {settings.tracker_url}")
    logger.info(f"  SMS_PROVIDER = {settings.sms_provider}")
    logger.info(f"  BROWSER_MODE = {settings.browser_mode}")

    if settings.sms_provider == "manual":
        if settings.manual_phone:
            logger.success(f"  ✓ телефон: {settings.manual_phone}")
        else:
            logger.info(
                "  ✓ телефон: задаётся в интерфейсе (Шаг 2) "
                "или вводится в консоли при python main.py run"
            )
    elif settings.sms_provider == "webhook":
        logger.success(
            f"  ✓ Android webhook: http://{settings.webhook_host}:"
            f"{settings.webhook_port}/sms"
        )
    else:
        logger.success("  ✓ телефон: через внешний SMS-провайдер")

    if settings.browser_mode == "cdp":
        try:
            ensure_cdp_chrome_running(
                port=CDP_PORT,
                use_default_profile=settings.cdp_use_default_profile,
            )
            logger.success(f"  ✓ Chrome CDP на порту {CDP_PORT}")
            logger.info(f"  BROWSER_LANDING = {settings.browser_landing}")
        except Exception as e:
            logger.error(f"  ✗ Chrome CDP: {e}")
            ok = False
        logger.info(
            "  ⚠ Web-Zaim часто даёт 403 через VPN. "
            "Для run выключи VPN или используй мобильный интернет."
        )

    logger.info(f"  таблица ТЗ: {settings.xlsx_path} (лист «sms»)")
    if ok:
        logger.success("Готово к запуску: python main.py run")
        return 0
    logger.error("Исправь .env и повтори: python main.py check")
    return 1


async def cmd_run() -> int:
    setup_logging()
    logger.info("=" * 72)
    logger.info("СТАРТ парсера МФО (Гидфинанс — тестовое + звонки)")
    logger.info(f"трекер: {settings.tracker_url}")
    logger.info(f"SMS-провайдер: {settings.sms_provider}  |  CALL-провайдер: {settings.call_provider}")
    logger.info(f"OpenAI: {'есть' if settings.openai_api_key else 'нет (rule-based)'}")
    logger.info("=" * 72)

    if settings.browser_mode == "cdp":
        ensure_cdp_chrome_running(
            port=CDP_PORT,
            use_default_profile=settings.cdp_use_default_profile,
        )

    effective_ua = resolve_user_agent(settings.user_agent)

    storage = EventStorage(settings.db_path, settings.csv_path, settings.xlsx_path)
    sms_provider = build_provider()
    call_provider = build_call_provider()
    ai = AIClassifier(api_key=settings.openai_api_key, model=settings.openai_model)

    rent = await sms_provider.rent_number()
    persona = generate_persona(email_domain=settings.fake_email_domain)
    persona.phone = rent.phone
    logger.info(f"персона: {persona.last_name} {persona.first_name} {persona.middle_name}")
    logger.info(f"телефон: {persona.phone}")

    run_state: dict[str, str] = {"landing_url": ""}
    screenshots_dir = settings.output_dir / "screenshots"

    code_event = threading.Event()
    code_holder: dict[str, str] = {}
    pending_code_holder: dict[str, str] = {}
    sms_input_ready = asyncio.Event()
    main_loop = asyncio.get_running_loop()

    def _signal_sms_input() -> None:
        def _mark_ready() -> None:
            sms_input_ready.set()
            pending_code = pending_code_holder.get("code")
            if pending_code and not code_event.is_set():
                logger.info(f"форма ждёт SMS: передаю сохранённый ранее код {pending_code}")
                code_holder["code"] = pending_code
                code_event.set()

        main_loop.call_soon_threadsafe(_mark_ready)

    form_task = asyncio.create_task(
        asyncio.to_thread(
            fill_and_submit_sync,
            tracker_url=settings.tracker_url,
            persona=persona,
            amount=settings.loan_amount,
            term_days=settings.loan_term_days,
            headless=settings.headless,
            screenshots_dir=screenshots_dir,
            browser_mode=settings.browser_mode,
            cdp_endpoint=settings.cdp_endpoint,
            user_agent=effective_ua,
            code_event=code_event,
            code_holder=code_holder,
            code_wait_seconds=settings.sms_listen_minutes * 60,
            allow_manual_recovery=settings.browser_manual_recovery,
            landing_mode=settings.browser_landing,
            use_default_profile=settings.cdp_use_default_profile,
            submit_enabled=not settings.dry_run_no_submit,
            keep_browser_open=settings.keep_browser_open,
            on_waiting_sms=_signal_sms_input,
        )
    )

    async def sms_loop() -> None:
        from src.form.sites.webzaim_selenium import extract_sms_code
        kwargs = (
            {"ready_event": sms_input_ready}
            if settings.sms_provider == "manual"
            else {}
        )
        async for sms in sms_provider.stream_sms(rent, **kwargs):
            caller = parse_manual_call_entry(sms.text)
            if caller:
                cls_call = await ai.classify_call(caller, 0)
                storage.log_call(
                    received_at=sms.received_at,
                    caller=caller,
                    duration_sec=0,
                    recording_url=None,
                    ai_category=cls_call.category,
                    ai_confidence=cls_call.confidence,
                    ai_summary=cls_call.summary,
                    risk_flag=cls_call.risk_flag,
                    mfo_landing=run_state["landing_url"],
                    phone=rent.phone,
                )
                logger.info("номер распознан как звонок, а не SMS")
                continue

            final, traces = await resolve_final_url(sms.text, user_agent=effective_ua)
            sender_alpha = infer_sender_alpha(sms.sender, sms.text, traces)
            cls = await ai.classify_sms(sender_alpha, sms.text)
            storage.log_sms(
                received_at=sms.received_at,
                sender_alpha=sender_alpha,
                text=sms.text,
                url_in_sms=traces[0].original if traces else None,
                final_redirect_url=final,
                redirect_hops=traces[0].hops if traces else None,
                ai_category=cls.category,
                ai_confidence=cls.confidence,
                ai_summary=cls.summary,
                risk_flag=cls.risk_flag,
                mfo_landing=run_state["landing_url"],
                phone=rent.phone,
                redirect_traces=traces,
            )
            code = extract_sms_code(sms.text)
            manual_code_only = (
                settings.sms_provider == "manual"
                and sms.text.strip().isdigit()
                and 4 <= len(sms.text.strip()) <= 6
            )
            otp_hint = bool(OTP_HINT_RX.search(f"{sender_alpha} {sms.text}"))
            webzaim_hint = bool(WEBZAIM_HINT_RX.search(f"{sender_alpha} {sms.text}"))
            form_is_waiting_for_code = sms_input_ready.is_set()
            looks_like_webzaim_otp = (
                webzaim_hint
                and (
                    cls.category == "verification_code"
                    or otp_hint
                )
            )
            should_autofill_code = (
                manual_code_only
                or (
                    form_is_waiting_for_code
                    and looks_like_webzaim_otp
                )
            )
            if (
                code
                and not code_event.is_set()
                and should_autofill_code
            ):
                logger.info(f"передаю код {code} в форму для автоввода")
                code_holder["code"] = code
                code_event.set()
            elif (
                code
                and not code_event.is_set()
                and looks_like_webzaim_otp
                and not form_is_waiting_for_code
            ):
                pending_code_holder["code"] = code
                logger.info(
                    f"код {code} пришёл раньше поля SMS — сохраняю и введу, "
                    "когда форма будет готова"
                )
            elif code and not code_event.is_set():
                logger.info(
                    f"нашёл похожие цифры {code}, но SMS не похожа на код "
                    f"подтверждения ({cls.category}) — не ввожу в форму"
                )

    async def call_loop() -> None:
        if not call_provider:
            return
        async for call in call_provider.stream_calls():
            cls = await ai.classify_call(call.caller, call.duration_sec)
            call_summary = " | ".join(
                part for part in (call.summary, cls.summary) if part
            )
            storage.log_call(
                received_at=call.started_at,
                caller=call.caller,
                duration_sec=call.duration_sec,
                recording_url=call.recording_url,
                ai_category=cls.category,
                ai_confidence=cls.confidence,
                ai_summary=call_summary,
                risk_flag=cls.risk_flag,
                mfo_landing=run_state["landing_url"],
                phone=rent.phone,
            )

    sms_task = asyncio.create_task(sms_loop())
    call_task = asyncio.create_task(call_loop())

    try:
        form_result = await form_task
        run_state["landing_url"] = form_result.landing_url
        if form_result.error:
            logger.error(f"форма: {form_result.error}")
        else:
            logger.success(f"форма отправлена: {form_result.submitted}")

        if not form_result.submitted:
            logger.warning(
                "форма не отправлена — SMS-код вряд ли придёт. "
                "Останавливаю SMS/call listener, чтобы не держать порт 8765."
            )
            await cancel_listener_tasks(sms_task, call_task)
            return 1
        elif settings.close_browser_after_submit:
            if settings.browser_mode == "cdp":
                logger.info("Chrome больше не нужен: закрываю браузер и продолжаю слушать SMS/звонки")
                kill_cdp_chrome(port=CDP_PORT)
            else:
                logger.info("Браузерный этап завершён, продолжаю слушать SMS/звонки")

        logger.info(
            f"теперь слушаю SMS + звонки {settings.sms_listen_minutes} мин на {rent.phone}…"
        )
        await asyncio.gather(sms_task, call_task)
    finally:
        await sms_provider.release(rent)
        if call_provider:
            await call_provider.close()
        await ai.close()
        storage.close()

    logger.success(f"готово. таблица ТЗ (лист «sms»): {settings.xlsx_path}")
    logger.info("  python main.py show   — посмотреть результат")
    return 0


async def cmd_demo(
    sample_text: str | None,
    no_browser: bool,
    with_calls: bool,
    sender: str,
    phone: str,
    landing: str,
) -> int:
    """Демо: парсим SMS, прогоняем редиректы, пишем в таблицу.

    Если --with-calls — дополнительно генерим 3 фейковых звонка через MockCallProvider.
    Параметрами --phone и --landing можно подставить реальные данные из своего
    ручного прогона: телефон-получатель и реальный лендинг МФО.
    """
    setup_logging()
    effective_ua = resolve_user_agent(settings.user_agent)
    storage = EventStorage(settings.db_path, settings.csv_path, settings.xlsx_path)
    ai = AIClassifier(api_key=settings.openai_api_key, model=settings.openai_model)

    sample_text = sample_text or (
        "CreditPlus: Vash kod 7421. "
        "Podrobnee: https://t.leads.tech/click/8/330/?sub1=bizdev&sub2=Name_vacancy"
    )
    logger.info(f"DEMO текст SMS от «{sender}» на {phone}: {sample_text}")
    urls = extract_urls(sample_text)
    logger.info(f"найдено URL: {urls}")

    final, traces = await resolve_final_url(
        sample_text, user_agent=effective_ua, use_browser=not no_browser
    )
    sender_alpha = infer_sender_alpha(sender, sample_text, traces)
    cls = await ai.classify_sms(sender_alpha, sample_text)
    storage.log_sms(
        received_at=datetime.now(),
        sender_alpha=sender_alpha,
        text=sample_text,
        url_in_sms=traces[0].original if traces else None,
        final_redirect_url=final,
        redirect_hops=traces[0].hops if traces else None,
        ai_category=cls.category,
        ai_confidence=cls.confidence,
        ai_summary=cls.summary,
        risk_flag=cls.risk_flag,
        mfo_landing=landing,
        phone=phone,
        redirect_traces=traces,
    )

    if with_calls:
        from src.calls.mock import MockCallProvider
        mock = MockCallProvider()
        logger.info("DEMO звонков: эмулирую 3 входящих звонка с задержками…")
        async for call in mock.stream_calls():
            cls_call = await ai.classify_call(call.caller, call.duration_sec)
            storage.log_call(
                received_at=call.started_at,
                caller=call.caller,
                duration_sec=call.duration_sec,
                recording_url=call.recording_url,
                ai_category=cls_call.category,
                ai_confidence=cls_call.confidence,
                ai_summary=cls_call.summary,
                risk_flag=cls_call.risk_flag,
                mfo_landing=landing,
                phone=phone,
            )

    await ai.close()
    storage.close()
    logger.success(f"строки добавлены в {settings.xlsx_path}")
    return 0


def _browser_demo_sync(keep_open: int, tracker_url: str, screenshots_dir: Path) -> str:
    screenshots_dir.mkdir(parents=True, exist_ok=True)
    screenshot_path = screenshots_dir / "browser_demo_landing.png"
    logger.info(f"browser-demo: режим {settings.browser_mode}")
    with open_browser(
        mode=settings.browser_mode,
        cdp_endpoint=settings.cdp_endpoint,
        headless=False,
        user_agent=resolve_user_agent(settings.user_agent),
    ) as driver:
        logger.info(f"Иду по партнёрской ссылке: {tracker_url}")
        driver.get(tracker_url)
        import time as _t
        _t.sleep(5)
        final = driver.current_url
        logger.success(f"Финальный лендинг: {final}")
        driver.save_screenshot(str(screenshot_path))
        logger.success(f"Скриншот сохранён: {screenshot_path}")
        logger.info(f"Оставляю окно открытым ещё {keep_open} сек…")
        _t.sleep(keep_open)
    return final


async def cmd_browser_demo(keep_open: int) -> int:
    """Открыть партнёрскую ссылку и показать редирект на лендинг МФО (без формы)."""
    setup_logging()
    screenshots_dir = settings.output_dir / "screenshots"
    await asyncio.to_thread(
        _browser_demo_sync, keep_open, settings.tracker_url, screenshots_dir
    )
    return 0


async def cmd_listen(landing: str, phone: str) -> int:
    """Тестовый режим: слушать SMS/звонки и писать их в таблицу без браузера."""
    setup_logging()
    logger.info("=" * 72)
    logger.info("СТАРТ webhook/listen: SMS + звонки → Excel без Selenium")
    logger.info(f"SMS-провайдер: {settings.sms_provider}  |  CALL-провайдер: {settings.call_provider}")
    logger.info(f"слушаю {settings.sms_listen_minutes} мин")
    logger.info("=" * 72)

    effective_ua = resolve_user_agent(settings.user_agent)
    storage = EventStorage(settings.db_path, settings.csv_path, settings.xlsx_path)
    sms_provider = build_provider()
    call_provider = build_call_provider()
    ai = AIClassifier(api_key=settings.openai_api_key, model=settings.openai_model)

    rent = await sms_provider.rent_number()
    if phone:
        rent.phone = phone
    run_landing = landing or "(listen)"

    async def sms_loop() -> None:
        async for sms in sms_provider.stream_sms(rent):
            caller = parse_manual_call_entry(sms.text)
            if caller:
                cls_call = await ai.classify_call(caller, 0)
                storage.log_call(
                    received_at=sms.received_at,
                    caller=caller,
                    duration_sec=0,
                    recording_url=None,
                    ai_category=cls_call.category,
                    ai_confidence=cls_call.confidence,
                    ai_summary=cls_call.summary,
                    risk_flag=cls_call.risk_flag,
                    mfo_landing=run_landing,
                    phone=rent.phone,
                )
                continue

            final, traces = await resolve_final_url(sms.text, user_agent=effective_ua)
            sender_alpha = infer_sender_alpha(sms.sender, sms.text, traces)
            cls = await ai.classify_sms(sender_alpha, sms.text)
            storage.log_sms(
                received_at=sms.received_at,
                sender_alpha=sender_alpha,
                text=sms.text,
                url_in_sms=traces[0].original if traces else None,
                final_redirect_url=final,
                redirect_hops=traces[0].hops if traces else None,
                ai_category=cls.category,
                ai_confidence=cls.confidence,
                ai_summary=cls.summary,
                risk_flag=cls.risk_flag,
                mfo_landing=run_landing,
                phone=rent.phone,
                redirect_traces=traces,
            )

    async def call_loop() -> None:
        if not call_provider:
            return
        async for call in call_provider.stream_calls():
            cls = await ai.classify_call(call.caller, call.duration_sec)
            call_summary = " | ".join(
                part for part in (call.summary, cls.summary) if part
            )
            storage.log_call(
                received_at=call.started_at,
                caller=call.caller,
                duration_sec=call.duration_sec,
                recording_url=call.recording_url,
                ai_category=cls.category,
                ai_confidence=cls.confidence,
                ai_summary=call_summary,
                risk_flag=cls.risk_flag,
                mfo_landing=run_landing,
                phone=rent.phone,
            )

    try:
        await asyncio.gather(sms_loop(), call_loop())
    finally:
        await sms_provider.release(rent)
        if call_provider:
            await call_provider.close()
        await ai.close()
        storage.close()

    logger.success(f"готово. данные записаны в {settings.xlsx_path}")
    return 0


async def cmd_log_call(caller: str, duration: int, landing: str, phone: str) -> int:
    """Залогировать реальный звонок с твоего телефона в таблицу."""
    setup_logging()
    storage = EventStorage(settings.db_path, settings.csv_path, settings.xlsx_path)
    ai = AIClassifier(api_key=settings.openai_api_key, model=settings.openai_model)
    cls = await ai.classify_call(caller, duration)
    storage.log_call(
        received_at=datetime.now(),
        caller=caller,
        duration_sec=duration,
        recording_url=None,
        ai_category=cls.category,
        ai_confidence=cls.confidence,
        ai_summary=cls.summary,
        risk_flag=cls.risk_flag,
        mfo_landing=landing,
        phone=phone or settings.manual_phone,
        redirect_traces=traces,
    )
    await ai.close()
    storage.close()
    logger.success(f"звонок {caller} ({duration}с) → {settings.xlsx_path}")
    return 0


async def cmd_log_sms(sender: str, text: str, landing: str, phone: str) -> int:
    """Залогировать одну SMS вручную (когда run уже закрыт)."""
    setup_logging()
    effective_ua = resolve_user_agent(settings.user_agent)
    storage = EventStorage(settings.db_path, settings.csv_path, settings.xlsx_path)
    ai = AIClassifier(api_key=settings.openai_api_key, model=settings.openai_model)
    final, traces = await resolve_final_url(text, user_agent=effective_ua, use_browser=False)
    sender_alpha = infer_sender_alpha(sender, text, traces)
    cls = await ai.classify_sms(sender_alpha, text)
    storage.log_sms(
        received_at=datetime.now(),
        sender_alpha=sender_alpha,
        text=text,
        url_in_sms=traces[0].original if traces else None,
        final_redirect_url=final,
        redirect_hops=traces[0].hops if traces else None,
        ai_category=cls.category,
        ai_confidence=cls.confidence,
        ai_summary=cls.summary,
        risk_flag=cls.risk_flag,
        mfo_landing=landing,
        phone=phone or settings.manual_phone,
        redirect_traces=traces,
    )
    await ai.close()
    storage.close()
    logger.success(f"SMS [{sender}] записана → {settings.xlsx_path}")
    return 0


async def cmd_show() -> int:
    setup_logging()
    if not settings.xlsx_path.exists():
        logger.warning("файла нет — сначала запустите run или demo")
        return 1
    from openpyxl import load_workbook
    wb = load_workbook(settings.xlsx_path)
    order = ["sms", "events", "calls"]
    names = [s for s in order if s in wb.sheetnames]
    names += [s for s in wb.sheetnames if s not in names]
    for sheet_name in names:
        print(f"\n=== лист «{sheet_name}»" + (" (ТЗ)" if sheet_name == "sms" else "") + " ===")
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            print(" | ".join("" if v is None else str(v)[:60] for v in row))
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Парсер МФО + SMS + звонки + AI")
    sub = parser.add_subparsers(dest="cmd", required=True)
    sub.add_parser("check", help="проверить .env, Chrome CDP, телефон")
    sub.add_parser("kill-chrome", help="закрыть Chrome, поднятый через CDP")
    sub.add_parser("run", help="полный сценарий: заявка + SMS + звонки")
    p_browser = sub.add_parser(
        "browser-demo",
        help="видимо открыть Chromium и показать редирект leads.tech → лендинг МФО",
    )
    p_browser.add_argument(
        "--keep-open",
        type=int,
        default=20,
        help="сколько секунд держать окно Chromium открытым",
    )
    p_listen = sub.add_parser(
        "listen",
        help="слушать SMS/звонки через выбранные провайдеры и писать в Excel без браузера",
    )
    p_listen.add_argument("--landing", default="(listen)", help="значение для колонки mfo_landing")
    p_listen.add_argument("--phone", default="", help="свой номер для колонки phone")
    p_demo = sub.add_parser("demo", help="демо без API SMS / без Zadarma")
    p_demo.add_argument("--text", help="произвольный текст SMS")
    p_demo.add_argument("--sender", default="WebZaim", help="альфа-имя отправителя SMS")
    p_demo.add_argument(
        "--phone",
        default="+79990000000",
        help="номер-получатель SMS (можно подставить свой реальный)",
    )
    p_demo.add_argument(
        "--landing",
        default="(demo)",
        help="реальный лендинг МФО для колонки mfo_landing",
    )
    p_demo.add_argument("--no-browser", action="store_true", help="без Chromium")
    p_demo.add_argument(
        "--with-calls",
        action="store_true",
        help="эмулировать входящие звонки через MockCallProvider",
    )
    sub.add_parser("show", help="вывести содержимое таблицы")
    sub.add_parser(
        "chrome",
        help="запустить Google Chrome с CDP (порт 9222) для режима BROWSER_MODE=cdp",
    )
    p_call = sub.add_parser("log-call", help="вручную залогировать пришедший звонок")
    p_call.add_argument("--caller", required=True, help="кто звонил, например +79837210053")
    p_call.add_argument("--duration", type=int, default=0, help="секунд")
    p_call.add_argument("--landing", default="", help="лендинг МФО")
    p_call.add_argument("--phone", default="", help="свой номер (по умолчанию из .env)")

    p_sms = sub.add_parser("log-sms", help="вручную залогировать одну SMS")
    p_sms.add_argument("--sender", default="web-zaim.ru")
    p_sms.add_argument("--text", required=True)
    p_sms.add_argument("--landing", default="")
    p_sms.add_argument("--phone", default="")
    args = parser.parse_args()

    if args.cmd == "check":
        return cmd_check()
    if args.cmd == "kill-chrome":
        setup_logging()
        kill_cdp_chrome(port=CDP_PORT)
        return 0
    if args.cmd == "chrome":
        ensure_cdp_chrome_running(
            use_default_profile=settings.cdp_use_default_profile,
        )
        logger.success(
            "Chrome запущен. Дальше: python main.py run\n"
            "  (в manual-режиме сначала открой ссылку в ЭТОМ Chrome, потом Enter)"
        )
        return 0
    if args.cmd == "run":
        return asyncio.run(cmd_run())
    if args.cmd == "browser-demo":
        return asyncio.run(cmd_browser_demo(args.keep_open))
    if args.cmd == "listen":
        return asyncio.run(cmd_listen(args.landing, args.phone))
    if args.cmd == "demo":
        return asyncio.run(
            cmd_demo(
                args.text,
                args.no_browser,
                args.with_calls,
                args.sender,
                args.phone,
                args.landing,
            )
        )
    if args.cmd == "show":
        return asyncio.run(cmd_show())
    if args.cmd == "log-call":
        return asyncio.run(
            cmd_log_call(args.caller, args.duration, args.landing, args.phone)
        )
    if args.cmd == "log-sms":
        return asyncio.run(
            cmd_log_sms(args.sender, args.text, args.landing, args.phone)
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
