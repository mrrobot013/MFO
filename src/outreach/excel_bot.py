"""Excel-only outreach workflow.

No messenger integration here: this module prepares messages and statuses in an
Excel file so MAX/SMS can be connected later without changing the decision logic.
"""
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

COLUMNS = [
    "id",
    "name",
    "phone",
    "max_id",
    "personal_url",
    "status",
    "first_message",
    "first_message_ready_at",
    "last_reply_text",
    "last_reply_at",
    "reply_class",
    "link_message",
    "link_ready_at",
    "notes",
]

POSITIVE_RX = re.compile(
    r"\b(да|давайте|давай|актуально|интересно|можно|скинь|пришл|хочу|ок|окей|ага|го|yes|yep)\b",
    re.IGNORECASE,
)
NEGATIVE_RX = re.compile(
    r"\b(нет|не надо|не актуально|неинтересно|отказ|стоп|stop|поздно|не хочу)\b",
    re.IGNORECASE,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


@dataclass
class ProcessResult:
    first_ready: int = 0
    link_ready: int = 0
    negative: int = 0
    unclear: int = 0
    errors: int = 0


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def normalize_name(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return ""
    return name.split()[0].strip().capitalize()


def first_message(name: str) -> str:
    first = normalize_name(name)
    if first:
        return f"Привет, {first}, ты хотел занять денег? Могу помочь с этим."
    return "Привет, ты хотел занять денег? Могу помочь с этим."


def link_message(personal_url: str) -> str:
    return f"Да, конечно. Вот твоя ссылка для оформления: {personal_url}"


def classify_reply(text: str) -> str:
    raw = (text or "").strip().lower()
    if not raw:
        return ""
    if NEGATIVE_RX.search(raw):
        return "negative"
    if POSITIVE_RX.search(raw):
        return "positive"
    return "unclear"


def create_template(path: Path, *, with_sample: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    widths = [8, 18, 18, 18, 55, 18, 55, 22, 45, 22, 16, 65, 22, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    if with_sample:
        ws.append([
            1,
            "Дима",
            "+79990000000",
            "",
            "https://example.com/personal-link-1",
            "new",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "пример строки, можно удалить",
        ])
    wb.save(path)


def ensure_workbook(path: Path) -> None:
    if not path.exists():
        create_template(path, with_sample=False)
        return
    wb = load_workbook(path)
    ws = wb["Leads"] if "Leads" in wb.sheetnames else wb.active
    existing = [cell.value for cell in ws[1]]
    changed = False
    for column in COLUMNS:
        if column not in existing:
            ws.cell(row=1, column=len(existing) + 1).value = column
            existing.append(column)
            changed = True
    if changed:
        wb.save(path)


def row_map(ws) -> dict[str, int]:
    return {str(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}


def get(ws, row: int, cols: dict[str, int], name: str) -> str:
    value = ws.cell(row=row, column=cols[name]).value
    return str(value or "").strip()


def set_value(ws, row: int, cols: dict[str, int], name: str, value: str) -> None:
    ws.cell(row=row, column=cols[name]).value = value


def process_workbook(path: Path, *, limit: int | None = None) -> ProcessResult:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb["Leads"] if "Leads" in wb.sheetnames else wb.active
    cols = row_map(ws)
    result = ProcessResult()
    processed = 0

    for row in range(2, ws.max_row + 1):
        if limit is not None and processed >= limit:
            break

        status = get(ws, row, cols, "status") or "new"
        name = get(ws, row, cols, "name")
        url = get(ws, row, cols, "personal_url")
        reply = get(ws, row, cols, "last_reply_text")

        if status == "new":
            set_value(ws, row, cols, "first_message", first_message(name))
            set_value(ws, row, cols, "first_message_ready_at", now())
            set_value(ws, row, cols, "status", "first_ready")
            result.first_ready += 1
            processed += 1
            continue

        if reply and status in {"first_ready", "first_sent", "replied_unclear"}:
            reply_class = classify_reply(reply)
            set_value(ws, row, cols, "reply_class", reply_class)
            if not get(ws, row, cols, "last_reply_at"):
                set_value(ws, row, cols, "last_reply_at", now())

            if reply_class == "positive":
                if not url:
                    set_value(ws, row, cols, "status", "error")
                    set_value(ws, row, cols, "notes", "нет personal_url")
                    result.errors += 1
                else:
                    set_value(ws, row, cols, "link_message", link_message(url))
                    set_value(ws, row, cols, "link_ready_at", now())
                    set_value(ws, row, cols, "status", "link_ready")
                    result.link_ready += 1
            elif reply_class == "negative":
                set_value(ws, row, cols, "status", "replied_negative")
                result.negative += 1
            else:
                set_value(ws, row, cols, "status", "replied_unclear")
                set_value(ws, row, cols, "notes", "нужно уточнение: актуально ли оформление")
                result.unclear += 1
            processed += 1

    wb.save(path)
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel outreach workflow without messenger integration")
    parser.add_argument("cmd", choices=["init", "process"])
    parser.add_argument("--file", default="data/leads.xlsx", help="path to leads workbook")
    parser.add_argument("--with-sample", action="store_true", help="add sample row when creating template")
    parser.add_argument("--limit", type=int, default=None, help="max rows to process")
    args = parser.parse_args()

    path = Path(args.file)
    if args.cmd == "init":
        create_template(path, with_sample=args.with_sample)
        print(f"created {path}")
        return 0

    result = process_workbook(path, limit=args.limit)
    print(
        "processed:",
        f"first_ready={result.first_ready}",
        f"link_ready={result.link_ready}",
        f"negative={result.negative}",
        f"unclear={result.unclear}",
        f"errors={result.errors}",
    )
    print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
