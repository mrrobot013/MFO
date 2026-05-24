"""Журнал событий: SMS + входящие звонки в SQLite + CSV + Excel.

В Excel два листа:
- `events` — все события в едином потоке (для аналитики timeline);
- `calls`  — только звонки с длительностью и ссылкой на запись.
- `sms`    — четыре обязательные колонки из ТЗ;
- `Links`  — ссылки из SMS и их финальные URL после редиректов;
- `Alphas` — сводка по отправителям.
- `Analytics` — итоговые счётчики и разрезы по дням/отправителям.
"""
from __future__ import annotations

import csv
import sqlite3
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Лист «sms» — ровно 4 колонки из ТЗ
TZ_SMS_COLUMNS = [
    "text",                 # текст сообщения
    "received_at",          # дата
    "sender_alpha",         # альфа-имя отправителя
    "final_redirect_url",   # финальная ссылка в редиректе
]

EVENT_COLUMNS = [
    "id",
    "event_type",          # sms | call
    "received_at",
    "sender_alpha",        # альфа-имя SMS или CallerID звонка
    "text",                # текст SMS или транскрипт звонка
    "duration_sec",        # только для звонков
    "recording_url",       # только для звонков
    "url_in_sms",
    "final_redirect_url",
    "redirect_hops",
    "ai_category",
    "ai_confidence",
    "ai_summary",
    "risk_flag",
    "mfo_landing",
    "phone",
]

CALL_COLUMNS = [
    "id",
    "received_at",
    "caller",
    "duration_sec",
    "recording_url",
    "ai_category",
    "ai_summary",
    "risk_flag",
    "phone",
]

LINK_COLUMNS = [
    "id",
    "sms_id",
    "sms_alpha",
    "sms_received",
    "tracker_type",
    "original_url",
    "final_url",
    "final_domain",
    "http_status",
    "redirect_chain",
]

ALPHA_COLUMNS = [
    "sms_alpha",
    "sms_count",
    "first_seen",
    "last_seen",
    "last_final_url",
]

ANALYTICS_METRIC_COLUMNS = ["metric", "value"]


SCHEMA_EVENTS = """
CREATE TABLE IF NOT EXISTS events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    event_type TEXT NOT NULL CHECK (event_type IN ('sms','call')),
    received_at TEXT NOT NULL,
    sender_alpha TEXT NOT NULL,
    text TEXT,
    duration_sec INTEGER,
    recording_url TEXT,
    url_in_sms TEXT,
    final_redirect_url TEXT,
    redirect_hops TEXT,
    ai_category TEXT,
    ai_confidence REAL,
    ai_summary TEXT,
    risk_flag INTEGER DEFAULT 0,
    mfo_landing TEXT,
    phone TEXT
);
"""

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
RISK_FILL = PatternFill("solid", fgColor="F4B084")


def _format_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _ensure_sheet(wb, name: str, columns: list[str], widths: list[int]):
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row == 0:
            ws.append(columns)
        existing = [cell.value for cell in ws[1]]
        if existing != columns:
            return ws
    else:
        ws = wb.create_sheet(name)
        ws.append(columns)
    _format_header(ws)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    return ws


def _set_sheet_order(wb) -> None:
    preferred = ["Links", "sms", "Alphas", "Applications", "calls", "events", "Analytics"]
    ordered = [wb[name] for name in preferred if name in wb.sheetnames]
    ordered += [ws for ws in wb.worksheets if ws.title not in preferred]
    wb._sheets = ordered


class EventStorage:
    def __init__(self, db_path: Path, csv_path: Path, xlsx_path: Path) -> None:
        self.db_path = db_path
        self.csv_path = csv_path
        self.xlsx_path = xlsx_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._conn = sqlite3.connect(self.db_path)
        self._conn.execute(SCHEMA_EVENTS)
        self._conn.commit()
        self._ensure_csv()
        self._ensure_xlsx()

    def _ensure_csv(self) -> None:
        if not self.csv_path.exists():
            with self.csv_path.open("w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(EVENT_COLUMNS)

    def _ensure_xlsx(self) -> None:
        if not self.xlsx_path.exists():
            wb = Workbook()
            wb.active.title = "events"
        else:
            wb = load_workbook(self.xlsx_path)

        _ensure_sheet(
            wb,
            "events",
            EVENT_COLUMNS,
            [5, 11, 20, 18, 60, 13, 35, 40, 50, 60, 18, 14, 50, 11, 40, 16],
        )
        _ensure_sheet(
            wb,
            "calls",
            CALL_COLUMNS,
            [5, 20, 22, 13, 40, 18, 50, 11, 16],
        )
        _ensure_sheet(
            wb,
            "sms",
            TZ_SMS_COLUMNS,
            [60, 20, 22, 55],
        )
        _ensure_sheet(
            wb,
            "Links",
            LINK_COLUMNS,
            [5, 8, 22, 20, 14, 55, 55, 24, 12, 90],
        )
        _ensure_sheet(
            wb,
            "Alphas",
            ALPHA_COLUMNS,
            [24, 12, 20, 20, 55],
        )
        if "Applications" not in wb.sheetnames:
            ws_app = wb.create_sheet("Applications")
            ws_app.append(["created_at", "phone", "mfo_landing", "comment"])
            _format_header(ws_app)
            for i, w in enumerate([20, 18, 60, 40], start=1):
                ws_app.column_dimensions[ws_app.cell(row=1, column=i).column_letter].width = w
        _set_sheet_order(wb)
        wb.save(self.xlsx_path)
        self._refresh_xlsx_analytics()

    def log_sms(
        self,
        received_at: datetime,
        sender_alpha: str,
        text: str,
        url_in_sms: str | None,
        final_redirect_url: str | None,
        redirect_hops: list[str] | None,
        ai_category: str,
        ai_confidence: float,
        ai_summary: str,
        risk_flag: bool,
        mfo_landing: str,
        phone: str,
        redirect_traces: list | None = None,
    ) -> int:
        hops_str = " → ".join(redirect_hops) if redirect_hops else ""
        cur = self._conn.execute(
            """INSERT INTO events
                  (event_type, received_at, sender_alpha, text,
                   url_in_sms, final_redirect_url, redirect_hops,
                   ai_category, ai_confidence, ai_summary, risk_flag,
                   mfo_landing, phone)
               VALUES ('sms', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                received_at.isoformat(timespec="seconds"),
                sender_alpha,
                text,
                url_in_sms,
                final_redirect_url,
                hops_str,
                ai_category,
                ai_confidence,
                ai_summary,
                1 if risk_flag else 0,
                mfo_landing,
                phone,
            ),
        )
        self._conn.commit()
        row_id = cur.lastrowid or 0

        row = [
            row_id, "sms", received_at.isoformat(timespec="seconds"),
            sender_alpha, text, "", "",
            url_in_sms or "", final_redirect_url or "", hops_str,
            ai_category, round(ai_confidence, 2), ai_summary,
            "⚠" if risk_flag else "", mfo_landing, phone,
        ]
        self._append_csv(row)
        self._append_xlsx_event(row, risk_flag=risk_flag)
        self._append_xlsx_tz_sms(
            text=text,
            received_at=received_at,
            sender_alpha=sender_alpha,
            final_redirect_url=final_redirect_url,
        )
        self._append_xlsx_links(row_id, received_at, sender_alpha, redirect_traces)
        self._refresh_xlsx_alphas()
        self._refresh_xlsx_analytics()

        logger.success(
            f"[SMS#{row_id}] {sender_alpha:>14}  [{ai_category}]  {text[:60]}"
            + (f"  →  {final_redirect_url}" if final_redirect_url else "")
        )
        return row_id

    def log_call(
        self,
        received_at: datetime,
        caller: str,
        duration_sec: int,
        recording_url: str | None,
        ai_category: str,
        ai_confidence: float,
        ai_summary: str,
        risk_flag: bool,
        mfo_landing: str,
        phone: str,
        transcript: str | None = None,
    ) -> int:
        cur = self._conn.execute(
            """INSERT INTO events
                  (event_type, received_at, sender_alpha, text,
                   duration_sec, recording_url,
                   ai_category, ai_confidence, ai_summary, risk_flag,
                   mfo_landing, phone)
               VALUES ('call', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                received_at.isoformat(timespec="seconds"),
                caller,
                transcript or "",
                duration_sec,
                recording_url,
                ai_category,
                ai_confidence,
                ai_summary,
                1 if risk_flag else 0,
                mfo_landing,
                phone,
            ),
        )
        self._conn.commit()
        row_id = cur.lastrowid or 0

        row_event = [
            row_id, "call", received_at.isoformat(timespec="seconds"),
            caller, transcript or "", duration_sec, recording_url or "",
            "", "", "",
            ai_category, round(ai_confidence, 2), ai_summary,
            "⚠" if risk_flag else "", mfo_landing, phone,
        ]
        row_call = [
            row_id, received_at.isoformat(timespec="seconds"), caller,
            duration_sec, recording_url or "",
            ai_category, ai_summary, "⚠" if risk_flag else "", phone,
        ]
        self._append_csv(row_event)
        self._append_xlsx_event(row_event, risk_flag=risk_flag)
        self._append_xlsx_calls(row_call, risk_flag=risk_flag)
        self._refresh_xlsx_analytics()

        duration_h = f"{duration_sec}s" if duration_sec else "no-answer"
        logger.success(
            f"[CALL#{row_id}] {caller:>14}  [{ai_category}]  {duration_h}  — {ai_summary[:60]}"
        )
        return row_id

    def _append_csv(self, row: list) -> None:
        with self.csv_path.open("a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(row)

    def _append_xlsx_event(self, row: list, risk_flag: bool) -> None:
        wb = load_workbook(self.xlsx_path)
        ws = wb["events"]
        ws.append(row)
        if risk_flag:
            for cell in ws[ws.max_row]:
                cell.fill = RISK_FILL
        wb.save(self.xlsx_path)

    def _append_xlsx_calls(self, row: list, risk_flag: bool) -> None:
        wb = load_workbook(self.xlsx_path)
        ws = wb["calls"]
        ws.append(row)
        if risk_flag:
            for cell in ws[ws.max_row]:
                cell.fill = RISK_FILL
        wb.save(self.xlsx_path)

    def _append_xlsx_tz_sms(
        self,
        *,
        text: str,
        received_at: datetime,
        sender_alpha: str,
        final_redirect_url: str | None,
    ) -> None:
        """Лист «sms» — только поля из ТЗ."""
        wb = load_workbook(self.xlsx_path)
        if "sms" not in wb.sheetnames:
            ws_tz = wb.create_sheet("sms")
            ws_tz.append(TZ_SMS_COLUMNS)
            _format_header(ws_tz)
        else:
            ws_tz = wb["sms"]
        ws_tz.append([
            text,
            received_at.isoformat(timespec="seconds"),
            sender_alpha,
            final_redirect_url or "",
        ])
        wb.save(self.xlsx_path)

    def _append_xlsx_links(
        self,
        sms_id: int,
        received_at: datetime,
        sender_alpha: str,
        traces: list | None,
    ) -> None:
        if not traces:
            return
        wb = load_workbook(self.xlsx_path)
        ws = _ensure_sheet(
            wb,
            "Links",
            LINK_COLUMNS,
            [5, 8, 22, 20, 14, 55, 55, 24, 12, 90],
        )
        next_id = ws.max_row
        for trace in traces:
            original = getattr(trace, "original", "") or ""
            final = getattr(trace, "final", "") or ""
            hops = getattr(trace, "hops", None) or []
            status_code = getattr(trace, "status_code", None)
            domain = urlparse(final).netloc.replace("www.", "") if final else ""
            ws.append([
                next_id,
                sms_id,
                sender_alpha,
                received_at.isoformat(timespec="seconds"),
                "requests",
                original,
                final,
                domain,
                status_code or "",
                " -> ".join(hops),
            ])
            next_id += 1
        wb.save(self.xlsx_path)

    def _refresh_xlsx_alphas(self) -> None:
        wb = load_workbook(self.xlsx_path)
        if "sms" not in wb.sheetnames:
            return
        ws_sms = wb["sms"]
        rows = list(ws_sms.iter_rows(min_row=2, values_only=True))
        stats: dict[str, dict[str, str | int]] = {}
        for text, received_at, sender_alpha, final_url in rows:
            sender = str(sender_alpha or "unknown")
            dt = str(received_at or "")
            item = stats.setdefault(
                sender,
                {
                    "sms_count": 0,
                    "first_seen": dt,
                    "last_seen": dt,
                    "last_final_url": "",
                },
            )
            item["sms_count"] = int(item["sms_count"]) + 1
            if dt and (not item["first_seen"] or dt < str(item["first_seen"])):
                item["first_seen"] = dt
            if dt and dt >= str(item["last_seen"]):
                item["last_seen"] = dt
                item["last_final_url"] = str(final_url or "")

        if "Alphas" in wb.sheetnames:
            del wb["Alphas"]
        ws_alpha = wb.create_sheet("Alphas")
        ws_alpha.append(ALPHA_COLUMNS)
        _format_header(ws_alpha)
        for i, w in enumerate([24, 12, 20, 20, 55], start=1):
            ws_alpha.column_dimensions[ws_alpha.cell(row=1, column=i).column_letter].width = w
        for sender, item in sorted(
            stats.items(),
            key=lambda kv: (-int(kv[1]["sms_count"]), kv[0].lower()),
        ):
            ws_alpha.append([
                sender,
                item["sms_count"],
                item["first_seen"],
                item["last_seen"],
                item["last_final_url"],
            ])
        wb.save(self.xlsx_path)

    def _refresh_xlsx_analytics(self) -> None:
        wb = load_workbook(self.xlsx_path)
        if "Analytics" in wb.sheetnames:
            del wb["Analytics"]
        ws = wb.create_sheet("Analytics")

        sms_rows = []
        if "sms" in wb.sheetnames:
            sms_rows = list(wb["sms"].iter_rows(min_row=2, values_only=True))
        call_rows = []
        if "calls" in wb.sheetnames:
            call_rows = list(wb["calls"].iter_rows(min_row=2, values_only=True))

        sms_by_day: dict[str, int] = {}
        calls_by_day: dict[str, int] = {}
        sender_stats: dict[str, dict[str, str | int]] = {}
        domain_stats: dict[str, int] = {}

        for text, received_at, sender_alpha, final_url in sms_rows:
            dt = str(received_at or "")
            day = dt[:10] if len(dt) >= 10 else "unknown"
            sender = str(sender_alpha or "unknown")
            final = str(final_url or "")
            domain = urlparse(final).netloc.replace("www.", "") if final else ""

            sms_by_day[day] = sms_by_day.get(day, 0) + 1
            if domain:
                domain_stats[domain] = domain_stats.get(domain, 0) + 1

            item = sender_stats.setdefault(
                sender,
                {
                    "sms_count": 0,
                    "first_seen": dt,
                    "last_seen": dt,
                    "last_final_url": "",
                },
            )
            item["sms_count"] = int(item["sms_count"]) + 1
            if dt and (not item["first_seen"] or dt < str(item["first_seen"])):
                item["first_seen"] = dt
            if dt and dt >= str(item["last_seen"]):
                item["last_seen"] = dt
                item["last_final_url"] = final

        for row in call_rows:
            received_at = row[1] if len(row) > 1 else ""
            dt = str(received_at or "")
            day = dt[:10] if len(dt) >= 10 else "unknown"
            calls_by_day[day] = calls_by_day.get(day, 0) + 1

        top_sender = ""
        if sender_stats:
            sender, item = max(
                sender_stats.items(),
                key=lambda kv: (int(kv[1]["sms_count"]), kv[0]),
            )
            top_sender = f"{sender} ({item['sms_count']})"

        top_domain = ""
        if domain_stats:
            domain, count = max(domain_stats.items(), key=lambda kv: (kv[1], kv[0]))
            top_domain = f"{domain} ({count})"

        def section(title: str, columns: list[str]) -> None:
            if ws.max_row > 1:
                ws.append([])
            ws.append([title])
            for cell in ws[ws.max_row]:
                cell.fill = HEAD_FILL
                cell.font = HEAD_FONT
            ws.append(columns)
            for cell in ws[ws.max_row]:
                cell.fill = HEAD_FILL
                cell.font = HEAD_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

        section("Summary", ANALYTICS_METRIC_COLUMNS)
        metrics = [
            ("total_sms", len(sms_rows)),
            ("total_calls", len(call_rows)),
            ("unique_sms_senders", len(sender_stats)),
            ("days_with_sms", len(sms_by_day)),
            ("top_sender", top_sender),
            ("top_final_domain", top_domain),
        ]
        for metric, value in metrics:
            ws.append([metric, value])

        section("SMS by day", ["date", "sms_count"])
        for day, count in sorted(sms_by_day.items()):
            ws.append([day, count])

        section("SMS by sender", ["sms_alpha", "sms_count", "first_seen", "last_seen", "last_final_url"])
        for sender, item in sorted(
            sender_stats.items(),
            key=lambda kv: (-int(kv[1]["sms_count"]), kv[0].lower()),
        ):
            ws.append([
                sender,
                item["sms_count"],
                item["first_seen"],
                item["last_seen"],
                item["last_final_url"],
            ])

        section("Final domains", ["final_domain", "sms_count"])
        for domain, count in sorted(domain_stats.items(), key=lambda kv: (-kv[1], kv[0])):
            ws.append([domain, count])

        section("Calls by day", ["date", "call_count"])
        for day, count in sorted(calls_by_day.items()):
            ws.append([day, count])

        for i, w in enumerate([24, 16, 22, 22, 70], start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        _set_sheet_order(wb)
        wb.save(self.xlsx_path)

    def close(self) -> None:
        self._conn.close()
