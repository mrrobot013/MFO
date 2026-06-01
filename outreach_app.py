from __future__ import annotations

from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook

from src.outreach.excel_bot import HEADERS, init_workbook, process_workbook


DEFAULT_FILE = Path("data/leads.xlsx")


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def read_leads(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=HEADERS)

    workbook = load_workbook(path)
    sheet = workbook["Leads"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=HEADERS)

    headers = [str(value) if value is not None else "" for value in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    frame = pd.DataFrame(data)
    for header in HEADERS:
        if header not in frame.columns:
            frame[header] = ""
    return frame[HEADERS]


def write_leads(path: Path, frame: pd.DataFrame) -> None:
    ensure_parent(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append(HEADERS)

    for _, row in frame.iterrows():
        sheet.append([row.get(header, "") for header in HEADERS])

    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)

    workbook.save(path)


def import_uploaded_file(uploaded_file, path: Path) -> None:
    ensure_parent(path)
    path.write_bytes(uploaded_file.getbuffer())


def export_bytes(frame: pd.DataFrame) -> bytes:
    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        write_leads(Path(tmp.name), frame)
        return Path(tmp.name).read_bytes()


def render_metrics(frame: pd.DataFrame) -> None:
    total = len(frame)
    first_ready = int((frame["status"] == "first_ready").sum()) if total else 0
    link_ready = int((frame["status"] == "link_ready").sum()) if total else 0
    positive = int((frame["reply_class"] == "positive").sum()) if total else 0

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Лидов", total)
    col2.metric("Первое сообщение", first_ready)
    col3.metric("Ссылка готова", link_ready)
    col4.metric("Позитивных ответов", positive)


def main() -> None:
    st.set_page_config(page_title="Обратка по лидам", page_icon="💬", layout="wide")

    st.title("💬 Обратка по лидам")
    st.caption("MVP без подключения к MAX/SMS: Excel-файл, ответы лидов, готовые персональные ссылки.")

    file_path = Path(st.sidebar.text_input("Файл лидов", str(DEFAULT_FILE)))
    uploaded_file = st.sidebar.file_uploader("Загрузить Excel", type=["xlsx"])

    if uploaded_file is not None:
        import_uploaded_file(uploaded_file, file_path)
        st.sidebar.success(f"Файл загружен: {file_path}")

    left, middle, right = st.sidebar.columns(3)
    if left.button("Создать", use_container_width=True):
        ensure_parent(file_path)
        init_workbook(file_path, with_sample=True)
        st.sidebar.success("Файл создан")

    if middle.button("Обработать", use_container_width=True):
        result = process_workbook(file_path)
        st.sidebar.success(
            "Готово: "
            f"first_ready={result.first_ready}, "
            f"link_ready={result.link_ready}, "
            f"negative={result.negative}, "
            f"unclear={result.unclear}, "
            f"errors={result.errors}"
        )

    if right.button("Обновить", use_container_width=True):
        st.rerun()

    frame = read_leads(file_path)
    render_metrics(frame)

    st.subheader("Лиды")
    st.caption(
        "Заполни `name`, `phone` или `max_id`, `personal_url`. "
        "Ответ лида вставляй в `last_reply_text`, затем нажимай «Сохранить и обработать»."
    )

    edited = st.data_editor(
        frame,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "personal_url": st.column_config.LinkColumn("personal_url"),
            "first_message": st.column_config.TextColumn("first_message", width="large"),
            "last_reply_text": st.column_config.TextColumn("last_reply_text", width="large"),
            "link_message": st.column_config.TextColumn("link_message", width="large"),
            "notes": st.column_config.TextColumn("notes", width="medium"),
        },
    )

    action1, action2, action3 = st.columns([1, 1, 2])
    if action1.button("Сохранить", type="primary", use_container_width=True):
        write_leads(file_path, edited)
        st.success(f"Сохранено: {file_path}")

    if action2.button("Сохранить и обработать", use_container_width=True):
        write_leads(file_path, edited)
        result = process_workbook(file_path)
        st.success(
            "Обработка завершена: "
            f"first_ready={result.first_ready}, "
            f"link_ready={result.link_ready}, "
            f"negative={result.negative}, "
            f"unclear={result.unclear}, "
            f"errors={result.errors}"
        )
        st.rerun()

    if not edited.empty:
        st.download_button(
            "Скачать leads.xlsx",
            data=export_bytes(edited),
            file_name=f"leads-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

    ready = edited[edited.get("status", "") == "link_ready"] if not edited.empty else pd.DataFrame()
    if not ready.empty:
        st.subheader("Готовые сообщения со ссылками")
        st.dataframe(
            ready[["name", "phone", "max_id", "last_reply_text", "link_message", "link_ready_at"]],
            use_container_width=True,
            hide_index=True,
        )


if __name__ == "__main__":
    main()
